"""
Rotas /financeiro — todos os endpoints financeiros principais.
"""

from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from openpyxl import load_workbook
from pydantic import BaseModel
from app.services.financeiro_service import FinanceiroService
from app.services.fluxo_caixa_service import FluxoCaixaService
from app.services.operacao_service import OperacaoService
from app.services.projection_service import ProjectionService
from app.services.saude_service import SaudeService

_EXCEL_PATH = Path(__file__).resolve().parents[3] / "data" / "Financeiro_BAM_Fase1_BI.xlsx"


class DespesaUpdate(BaseModel):
    valor: Optional[float] = None
    descricao: Optional[str] = None
    status: Optional[str] = None
    categoria: Optional[str] = None


class ReceitaUpdate(BaseModel):
    valor: Optional[float] = None
    descricao: Optional[str] = None
    status: Optional[str] = None


def _update_excel_row(sheet_name: str, row_excel: int, updates: dict) -> None:
    """Abre o Excel, atualiza a linha indicada e salva."""
    wb = load_workbook(_EXCEL_PATH, keep_links=False)
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    for field, value in updates.items():
        if field in col_map and value is not None:
            ws.cell(row=row_excel, column=col_map[field]).value = value
    wb.save(_EXCEL_PATH)
    wb.close()

router = APIRouter()
_fin   = FinanceiroService()
_proj  = ProjectionService()
_saude = SaudeService()
_ops   = OperacaoService()
_fluxo = FluxoCaixaService()


class CaixaUpdateRequest(BaseModel):
    valor_atual: float
    observacao: str = ""


@router.get("/resumo")
def get_resumo():
    """Visão consolidada mensal de receitas e despesas."""
    try:
        return _fin.get_resumo_mensal()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/kpis")
def get_kpis():
    """KPIs principais: receita total, despesa, resultado, margem, lançamentos."""
    try:
        return _fin.get_kpis()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/receitas")
def get_receitas():
    """Lista e agregações de receitas."""
    try:
        return _fin.get_receitas()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/despesas")
def get_despesas():
    """Lista e agregações de despesas por categoria."""
    try:
        return _fin.get_despesas()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.put("/despesas/{id}")
def update_despesa(id: int, body: DespesaUpdate):
    """Atualiza uma despesa diretamente no Excel. id = índice do lançamento (1-based)."""
    if not _EXCEL_PATH.exists():
        raise HTTPException(status_code=503, detail="Arquivo Excel não encontrado")
    # pandas row index id-1 → excel row = id + 1 (header na linha 1)
    excel_row = id + 1
    mapping = {
        "Valor":     body.valor,
        "Despesa":   body.descricao,
        "Status":    body.status,
        "Categoria": body.categoria,
    }
    updates = {k: v for k, v in mapping.items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    try:
        _update_excel_row("Base Despesas", excel_row, updates)
        return {"ok": True, "row": excel_row, "updates": updates}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.put("/receitas/{id}")
def update_receita(id: int, body: ReceitaUpdate):
    """Atualiza uma receita diretamente no Excel. id = índice do lançamento (1-based)."""
    if not _EXCEL_PATH.exists():
        raise HTTPException(status_code=503, detail="Arquivo Excel não encontrado")
    excel_row = id + 1
    mapping = {
        "Valor Previsto": body.valor,
        "Descrição":      body.descricao,
        "Status":         body.status,
    }
    updates = {k: v for k, v in mapping.items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    try:
        _update_excel_row("Base Receitas", excel_row, updates)
        return {"ok": True, "row": excel_row, "updates": updates}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/projecoes")
def get_projecoes(
    crescimento_pct: float = Query(default=10.0, description="Crescimento percentual mensal de receita"),
    novos_clientes_crm: int = Query(default=10, description="Novos clientes no módulo CRM (R$100/cliente)"),
    meses: int = Query(default=6, ge=1, le=24, description="Horizonte de projeção em meses"),
):
    """Simulação de crescimento com parâmetros configuráveis."""
    try:
        return _proj.get_projecoes(
            crescimento_pct=crescimento_pct,
            novos_clientes_crm=novos_clientes_crm,
            meses=meses,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/cenarios")
def get_cenarios():
    """Cenários de corte de despesas: conservador, moderado e agressivo."""
    try:
        return _proj.get_cenarios_corte()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/alertas")
def get_alertas():
    """Alertas financeiros automáticos baseados nos KPIs."""
    try:
        return _fin.get_alertas()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/saude")
def get_saude():
    """Score de saúde financeira, semáforo, riscos e recomendações."""
    try:
        return _saude.get_saude()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/insights")
def get_insights():
    """Insights automáticos baseados nos dados financeiros."""
    try:
        return _saude.get_insights()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/operacao-mes")
def get_operacao_mes():
    """Visão operacional do mês com receber x pagar, caixa e conferência diária."""
    try:
        return _ops.get_operacao_mes()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/caixa")
def get_caixa():
    """Retorna caixa atual, caixa anterior e histórico de atualizações."""
    try:
        return _ops.get_caixa()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/caixa")
def update_caixa(body: CaixaUpdateRequest):
    """Atualiza manualmente o caixa atual com histórico."""
    try:
        registro = _ops.update_caixa(
            valor_atual=body.valor_atual,
            observacao=body.observacao,
            origem="manual",
        )
        return {"ok": True, "registro": registro}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ── Fluxo de Caixa ───────────────────────────────────────────────────────

@router.get("/fluxo-caixa")
def get_fluxo_caixa(
    mes:       Optional[int] = Query(default=None, description="Mês (1–12)"),
    ano:       Optional[int] = Query(default=None, description="Ano ex: 2026"),
    tipo:      Optional[str] = Query(default=None, description="entrada | saida"),
    status:    Optional[str] = Query(default=None, description="previsto | recebido | pago | vencido"),
    cliente:   Optional[str] = Query(default=None, description="Filtro parcial de nome de cliente"),
    categoria: Optional[str] = Query(default=None, description="Filtro parcial de categoria"),
):
    """Fluxo de caixa unificado (Excel + movimentações manuais), com totais e KPIs."""
    try:
        return _fluxo.get_fluxo(
            mes=mes, ano=ano, tipo=tipo,
            status=status, cliente=cliente, categoria=categoria,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ── Conciliação ──────────────────────────────────────────────────────────

@router.get("/conciliacao")
def get_conciliacao(
    mes: Optional[int] = Query(default=None),
    ano: Optional[int] = Query(default=None),
):
    """KPIs e lançamentos para conferência/conciliação com extrato."""
    try:
        return _fluxo.get_conciliacao(mes=mes, ano=ano)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class ConciliacaoMarcarRequest(BaseModel):
    lancamento_id:      str
    status_conciliacao: str          # conciliado | pendente | divergente
    observacao:         str   = ""
    valor_extrato:      Optional[float] = None


@router.post("/conciliacao/marcar")
def marcar_conciliacao(body: ConciliacaoMarcarRequest):
    """Marca um lançamento como conciliado, pendente ou divergente."""
    try:
        return _fluxo.marcar_conciliacao(
            lancamento_id=body.lancamento_id,
            status_conciliacao=body.status_conciliacao,
            observacao=body.observacao,
            valor_extrato=body.valor_extrato,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ── Atualização manual de status de lançamento ──────────────────────────

class LancamentoStatusRequest(BaseModel):
    lancamento_id:   str
    status:          str            # recebido | previsto | pago | cancelado
    valor_realizado: Optional[float] = None


@router.post("/lancamento/status")
def update_lancamento_status(body: LancamentoStatusRequest):
    """Altera manualmente o status de um lançamento (override sobre o Excel)."""
    STATUS_VALIDOS = {"recebido", "previsto", "pago", "cancelado", "pendente", "vencido"}
    if body.status not in STATUS_VALIDOS:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {STATUS_VALIDOS}")
    try:
        return _fluxo.update_status(
            lancamento_id=body.lancamento_id,
            status=body.status,
            valor_realizado=body.valor_realizado,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# ── Recebimentos por Cliente ─────────────────────────────────────────────

@router.get("/recebimentos-clientes")
def get_recebimentos_clientes(
    mes: Optional[int] = Query(default=None),
    ano: Optional[int] = Query(default=None),
):
    """Visão consolidada de recebimentos por cliente ativo no mês."""
    try:
        return _fluxo.get_recebimentos_clientes(mes=mes, ano=ano)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
